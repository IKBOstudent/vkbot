import re
import requests
from bs4 import BeautifulSoup
import openpyxl
import datetime


def find_group(group):
    group = group.upper()
    found = False

    page = requests.get("https://www.mirea.ru/schedule/")
    soup = BeautifulSoup(page.text, "html.parser")

    result = soup.find("div", {"class": "rasspisanie"}). \
        find(string="Институт информационных технологий"). \
        find_parent("div").find_parent("div").find_all("a", {"class": "uk-link-toggle"})

    course = datetime.datetime.now().year % 100 - int(group[-2:])
    if 1 <= course <= 4:
        for a_tag in result:
            if re.search(r"ИИТ_" + str(course), a_tag['href']):
                table = requests.get(a_tag['href'])

                with open("table.xlsx", "wb") as f:
                    f.write(table.content)

                book = openpyxl.load_workbook("table.xlsx")
                sheet = book.active
                num_cols = sheet.max_column

                for col in range(1, num_cols):
                    if str(sheet.cell(row=2, column=col).value) == group:
                        found = True
                        break
    return found
