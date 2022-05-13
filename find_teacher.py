import re
import requests
from bs4 import BeautifulSoup
import openpyxl
import datetime


def find_teacher(teacher):
    teacher = teacher.lower()
    different_teachers = []

    page = requests.get("https://www.mirea.ru/schedule/")
    soup = BeautifulSoup(page.text, "html.parser")

    result = soup.find("div", {"class": "rasspisanie"}). \
        find(string="Институт информационных технологий"). \
        find_parent("div").find_parent("div").find_all("a", {"class": "uk-link-toggle"})

    for a_tag in result:
        table = requests.get(a_tag['href'])

        with open("table.xlsx", "wb") as f:
            f.write(table.content)

        book = openpyxl.load_workbook("table.xlsx")
        sheet = book.active
        num_cols = sheet.max_column
        num_rows = sheet.max_row

        for col in range(1, num_cols):
            if re.search(r"[а-яА-Я]{4}\-\d\d\-\d\d", str(sheet.cell(row=2, column=col).value)):
                for row in range(4, num_rows):
                    cell_teacher = sheet.cell(row=row, column=col + 2).value
                    if cell_teacher:
                        cell_teacher = str(cell_teacher)
                        searched = re.findall(r"[А-Яа-яЁё]+-?[А-Яа-яЁё]+  ?[А-Яа-я]?[. ,]*[А-Яа-я]?\.*", cell_teacher)
                        if searched:
                            for i in searched:
                                if i.find("  ") != -1:
                                    i = i.replace("  ", " ")
                                if i.find("..") != -1:
                                    i = i.replace("..", ".")
                                i = i.rstrip()

                                if i == "Иоффе Н Е." or i == "Ануфриев О. С.":
                                    continue
                                if i[-1] != "." and i != "Гриценко":
                                    i += "."
                                if teacher in i.lower() and i not in different_teachers:
                                    different_teachers.append(i)
                                    if len(different_teachers) > 10:
                                        return different_teachers

    return different_teachers
