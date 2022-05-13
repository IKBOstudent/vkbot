import re
import os
import requests
from bs4 import BeautifulSoup
import openpyxl
import math
import datetime
import json

weekdays = ["ПН", "ВТ", "СР", "ЧТ", "ПТ", "СБ"]


def parser_for_weeks(s, flag_krome):
    if flag_krome:
        s = s.replace("кр.", "")
    s = s.replace("н.", "")
    s = s.strip()
    weeks = s.split(',')
    for w in range(len(weeks)):
        if "-" in weeks[w]:
            new = weeks[w].split('-')
            weeks[w] = new[0]
            for j in range(int(new[0]) + 1, int(new[1]) + 1):
                weeks.append(j)

    weeks = list(map(int, weeks))
    return weeks


def formatted_message(file_data, weekday, col_name, current_week):
    message = ""
    for i in range(1, 7):
        num = str(i)
        subject = file_data["timetable"][col_name][weekdays[weekday]][num]["subject"]
        type_sub = file_data["timetable"][col_name][weekdays[weekday]][num]["type"]
        aud = file_data["timetable"][col_name][weekdays[weekday]][num]["aud"]
        teacher = file_data["timetable"][col_name][weekdays[weekday]][num]["teacher"]

        if subject == "-":
            message += f"\n{num}. -"
        else:
            subject = subject.replace("  ", " ")
            count = 0
            chose = -1
            no_split = False
            no_split_subject = []
            if "(1 п/г)" in subject:
                no_split = True

            for sub in subject.split("\n"):
                count += 1

                good = True
                match2 = re.findall(r"кр\. [0-9].*н\. ", sub)
                sub = re.sub(r"кр\. [0-9].*н\. ", "", sub).strip()

                match1 = re.findall(r"[0-9].*н\. ", sub)
                sub = re.sub(r"[0-9].*н\. ", "", sub).strip()
                if match2:
                    for match in match2:
                        weeks = parser_for_weeks(match, True)
                        if current_week in weeks:
                            good = False

                elif match1:
                    for match in match1:
                        weeks = parser_for_weeks(match, False)
                        if current_week not in weeks:
                            good = False

                if good:
                    if no_split:
                        chose = 0
                        no_split_subject.append(sub)
                    else:
                        chose = count - 1
                        subject = sub
                        break

            if no_split:
                subject = " | ".join(no_split_subject)

            if chose != -1:
                message += "\n" + f'{i}. {subject}'
                if type_sub != '-':
                    if not no_split:
                        type_sub = type_sub.split('\n')[chose]
                    else:
                        type_sub = type_sub.replace("\n", " | ")
                    type_sub = type_sub.replace("  ", " ")
                    message += f' ({type_sub})'
                if aud != '-':
                    if not no_split:
                        aud = aud.split('\n')[chose]
                    else:
                        aud = aud.replace("\n", " | ")
                    aud = aud.replace("  ", " ")
                    message += f', ауд. {aud}'
                if teacher != '-':
                    if not no_split:
                        teacher = teacher.split('\n')[chose]
                    else:
                        teacher = teacher.replace("\n", " | ")
                    teacher = teacher.replace("  ", " ")
                    message += f', преп. {teacher}'
            else:
                message += f"\n{num}. -"

    return message


def make_group_schedule_message(group, command):
    group = group.upper()
    print("group", group, command)
    current_week = datetime.datetime.now().date() - \
                   datetime.datetime.strptime("07.02.2022", '%d.%m.%Y').date()
    current_week = math.ceil(current_week.days / 7)
    current_week = 16
    even_week = current_week % 2 == 0

    today = datetime.datetime.now().weekday()
    tomorrow = (datetime.datetime.now() + datetime.timedelta(days=1)).weekday()

    # print("week:", current_week,
    #       "today:", weekdays[today],
    #       "tomorrow:", weekdays[tomorrow])

    page = requests.get("https://www.mirea.ru/schedule/")
    soup = BeautifulSoup(page.text, "html.parser")

    result = soup.find("div", {"class": "rasspisanie"}). \
        find(string="Институт информационных технологий"). \
        find_parent("div").find_parent("div").find_all("a", {"class": "uk-link-toggle"})

    file_data = {"group": group, "timetable": {}}

    course = datetime.datetime.now().year % 100 - int(group[-2:])
    for a_tag in result:
        if re.search(r"ИИТ_" + str(course), a_tag['href']):
            table = requests.get(a_tag['href'])
            print(table)

            with open("table.xlsx", "wb") as f:
                f.write(table.content)

            book = openpyxl.load_workbook("table.xlsx")
            sheet = book.active
            num_cols = sheet.max_column
            num_rows = sheet.max_row

            find_group_col = 0
            for col in range(1, num_cols):
                if str(sheet.cell(row=2, column=col).value) == group:
                    find_group_col = col
                    break

            schedule_data = {"odd": {}, "even": {}}
            for d in weekdays:
                schedule_data["odd"].update({d: {}})
                schedule_data["even"].update({d: {}})
                for n in "123456":
                    schedule_data["odd"][d].update({n: {}})
                    schedule_data["even"][d].update({n: {}})
                    for f in "subject", "teacher", "type", "aud":
                        schedule_data["odd"][d][n].update({f: '-'})
                        schedule_data["even"][d][n].update({f: '-'})

            i = 0
            for row in range(4, 76):
                subject = sheet.cell(row=row, column=find_group_col).value
                sub_type = sheet.cell(row=row, column=find_group_col + 1).value
                teacher = sheet.cell(row=row, column=find_group_col + 2).value
                aud = sheet.cell(row=row, column=find_group_col + 3).value

                if (row - 4) % 12 == 0:
                    if row != 4:
                        i += 1

                if row % 2 == 0:
                    num = str((((row - 4) % 12) + 2) // 2)
                    col_name = "odd"
                else:
                    num = str((((row - 4) % 12) + 1) // 2)
                    col_name = "even"

                if re.search(r"[А-Яа-я]+", str(subject)):
                    schedule_data[col_name][weekdays[i]][num]["subject"] = str(subject)
                    if teacher and not re.search(r"[0-9]", str(teacher)):
                        schedule_data[col_name][weekdays[i]][num]["teacher"] = str(teacher)
                    if sub_type:
                        schedule_data[col_name][weekdays[i]][num]["type"] = str(sub_type)
                    if aud:
                        schedule_data[col_name][weekdays[i]][num]["aud"] = str(aud)

            file_data["timetable"] = schedule_data

    filepath = os.path.abspath(os.curdir) + "/tables/group_schedule.json"
    with open(filepath, 'w', encoding="utf-8") as file:
        json.dump(file_data, file, indent=4, ensure_ascii=False)

    if command == "TOD":
        message = "Расписание на сегодня"
        col_name = "even" if even_week else "odd"
        message += formatted_message(file_data, today, col_name, current_week)
        print(message)

    elif command == "TOM":
        message = "Расписание на завтра"
        col_name = "even" if even_week else "odd"
        if True or tomorrow == 0:
            current_week += 1
            col_name = "odd" if even_week else "even"
        message += formatted_message(file_data, tomorrow, col_name, current_week)

        print(message)

    elif command == "THIS WEEK":
        message = "Расписание на эту неделю"
        col_name = "even" if even_week else "odd"
        for i in range(6):
            message += "\n" + formatted_message(file_data, i, col_name, current_week)
        print(message)

    elif command == "NEXT WEEK":
        message = "Расписание на следующую неделю"
        col_name = "odd" if even_week else "even"
        for i in range(6):
            message += "\n" + formatted_message(file_data, i, col_name, current_week+1)
        print(message)
    else:  # command - день недели
        pass


