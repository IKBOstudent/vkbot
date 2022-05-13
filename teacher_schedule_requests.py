import re
import requests
from bs4 import BeautifulSoup
import openpyxl
import datetime
import json
import os
import math


def make_teacher_schedule_message(teacher, command):
    print("teacher", teacher, command)
    teacher = teacher.lower()
    if teacher[-1] == ".":
        teacher = teacher[:-1]

    current_week = datetime.datetime.now().date() - \
                   datetime.datetime.strptime("07.02.2022", '%d.%m.%Y').date()
    current_week = math.ceil(current_week.days / 7)
    even_week = current_week % 2 == 0

    weekdays = ["ПН", "ВТ", "СР", "ЧТ", "ПТ", "СБ"]
    today = datetime.datetime.now().weekday()
    tomorrow = (datetime.datetime.now() + datetime.timedelta(days=1)).weekday()

    # print("week:", current_week,
    #       "today:", weekdays[today],
    #       "tomorrow:", weekdays[tomorrow])

    page = requests.get("https://www.mirea.ru/schedule/")
    soup = BeautifulSoup(page.text, "html.parser")

    result = soup.find("div", {"class": "rasspisanie"}). \
        find(string="Институт информационных технологий"). \
        find_parent("div").find_parent("div").find_all("a", {"class": "uk-link-toggle"})

    file_data = {"group": teacher, "timetable": {}}

    schedule_data = {"odd": {}, "even": {}}
    for d in weekdays:
        schedule_data["odd"].update({d: {}})
        schedule_data["even"].update({d: {}})
        for n in "123456":
            schedule_data["odd"][d].update({n: {}})
            schedule_data["even"][d].update({n: {}})
            for f in "group", "subject", "aud":
                schedule_data["odd"][d][n].update({f: '-'})
                schedule_data["even"][d][n].update({f: '-'})

    for a_tag in result:
        table = requests.get(a_tag['href'])

        with open("table.xlsx", "wb") as f:
            f.write(table.content)

        book = openpyxl.load_workbook("table.xlsx")
        sheet = book.active
        num_cols = sheet.max_column
        num_rows = sheet.max_row

        for col in range(1, num_cols):
            if re.search(r"[а-яА-Я]{4}\-\d\d\-\d\d", str(sheet.cell(row=2, column=col).value)):
                i = 0
                for row in range(4, num_rows):
                    if (row - 4) % 12 == 0:
                        if row != 4:
                            i += 1
                    cell_teacher = sheet.cell(row=row, column=col + 2).value
                    if cell_teacher:
                        if teacher in str(cell_teacher).lower():

                            if row % 2 == 0:
                                num = str((((row - 4) % 12) + 2) // 2)
                                col_name = "odd"
                            else:
                                num = str((((row - 4) % 12) + 1) // 2)
                                col_name = "even"

                            # print(str(cell_teacher))
                            group = str(sheet.cell(row=2, column=col).value)
                            subject = str(sheet.cell(row=row, column=col).value)
                            aud = str(sheet.cell(row=row, column=col+3).value)

                            schedule_data[col_name][weekdays[i]][num]["group"] = group
                            schedule_data[col_name][weekdays[i]][num]["subject"] = subject
                            schedule_data[col_name][weekdays[i]][num]["aud"] = aud

    file_data["timetable"] = schedule_data

    filepath = os.path.abspath(os.curdir) + "/tables/teacher_schedule.json"
    with open(filepath, 'w', encoding="utf-8") as file:
        json.dump(file_data, file, indent=4, ensure_ascii=False)
    #
    # with open(filepath, 'r', encoding="utf-8") as file:
    #     schedule = json.load(file)

